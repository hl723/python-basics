#! python3
# multiplication_table.py - creates a multiplication table in excel with N rows
"""
Create a program multiplicationTable.py that takes a number N from the command line
and creates an N×N multiplication table in an Excel spreadsheet.

Note with this module! There have been updates to openpyxl since Automate The Boring Stuff
was written and it has caused a lot of syntax to change.
"""

import sys
import openpyxl
import os
# Notice, we have to use openpyxyl.utils instead of openpyxl.cell now
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# sys.argv[1] will be the first argument entered into the command line
n_table = int(sys.argv[1])

# Open a new workbook and make your sheet active.
wb = openpyxl.Workbook()
sheet = wb.active

# Compute the headings and set them to bold
for r in range(2, n_table + 2):
    sheet['A' + str(r)] = r - 1
    col_let = get_column_letter(r)
    sheet[col_let + str(1)] = r - 1
    # Set font to bold
    sheet['A' + str(r)].font = Font(bold=True)
    sheet[col_let + str(1)].font = Font(bold=True)

# Now populate the table
for col in range(2, n_table + 2):
    for row in range(2, n_table + 2):
        col_letters = get_column_letter(col)
        sheet[col_letters + str(row)] = (col - 1) * (row - 1)

wb.save('Multiplication Table.xlsx')
